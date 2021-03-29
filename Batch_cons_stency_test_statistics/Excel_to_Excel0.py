#!/usr/bin/python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import numpy as np
import os
import re
import time
import json
from xlwt import *
import pandas as pd

'''
#获取真实的有效行数
def get_max_row(sheet):
    i=ws.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break

    return real_max_row
'''

if __name__ == '__main__':
    start = time.time()

    #设置数据文件夹路径
    path = ".\\data\\"
    final_data = {}
    #获取目录下的所有xlsx文件，并存入列表中
    xlsx_list = os.listdir(path)    #不是按顺序读取的
    xlsx_list.sort(key = lambda x : int(x.split('.')[0][8:])) #使用sort进行按顺序读取
    xlsx_num = 0
    #获取芯片编号
    chip_num = re.split("[_.]", xlsx_list[xlsx_num])[-2]
    for xlsx_num in range(len(xlsx_list)):
        #获取芯片编号
        chip_num = re.split("[_.]", xlsx_list[xlsx_num])[-2]
        #加载xlsx文件
        xlsx_name = str(path + xlsx_list[xlsx_num])
        wb = load_workbook(xlsx_name)
        #读取该workbook下的所有表格
        sheetnames = wb.sheetnames
        #定义表格索引值
        n = 0
        #内层循环，提取单个xlsx文件中各个表的统计值
        for n in range(len(sheetnames)):
            #指定表名
            ws = wb[wb.sheetnames[n]]
            '''
            #删除第 5 列数据，即STEP列
            ws.delete_cols(5)
            #删除第 1 列数据，即ACC列
            ws.delete_cols(1)
            
            #获取最大行、列
            max_row_value = ws.max_row
            max_column_value = ws.max_column

            #删除第最后一行以及第一行数据
            ws.delete_rows(max_row_value)
            ws.delete_rows(2)
            print(max_row_value)
            '''
            max_row_value = ws.max_row
            max_column_value = ws.max_column

            #print(max_row_value)

            #获取某个区间的值，例：获得了以A1为左上角，C90为右下角矩形区域的所有单元格
            #data_list即为所提取出的XYZ列表值
            data_list = []
            offset_list = []
            #for row_cell in ws['A2':'C80' ]:
            for row_cell in ws['B3':'D'+str(max_row_value-1)]:
                every_row_value = []
                for cell in row_cell:

                    every_row_value.append(cell.value)
                    #将列表转换为float类型
                    every_row_value = list(map(float, every_row_value))
                data_list.append(every_row_value)
            #print(data_list)
                
            #计算XYZ的均值
            mean_xyz_list = []
            data_list = np.array(data_list)
            data_list = data_list.astype(np.float64)
            mean_x = np.mean(data_list[:, 0])
            mean_y = np.mean(data_list[:, 1])
            mean_z = np.mean(data_list[:, 2])
            mean_xyz_list = [mean_x, mean_y, mean_z]


            #计算XYZ的最大值与最小值
            max_xyz_list = []
            max_x = np.max(data_list[:, 0])
            max_y = np.max(data_list[:, 1])
            max_z = np.max(data_list[:, 2])
            max_xyz_list = [max_x, max_y, max_z]

            min_xyz_list = []
            min_x = np.min(data_list[:, 0])
            min_y = np.min(data_list[:, 1])
            min_z = np.min(data_list[:, 2])
            min_xyz_list = [min_x, min_y, min_z]

            #计算XYZ均方差
            std_xyz_list = []
            std_x = np.std(data_list[:, 0])
            std_y = np.std(data_list[:, 1])
            std_z = np.std(data_list[:, 2])
            std_xyz_list = [std_x, std_y, std_z]
            std_xyz_list = list((np.array(std_xyz_list)/9.807)*1000)


            #计算offset值
            #设置各姿态下的加速度标准值
            theo_value_x1 = [9.8, 0, 0]
            theo_value_x0 = [-9.8, 0, 0]
            theo_value_y1 = [0, 9.8, 0]
            theo_value_y0 = [0, -9.8, 0]
            theo_value_z1 = [0, 0, 9.8]
            theo_value_z0 = [0, 0, -9.8]
            #根据当前的n值确定计算方法
            if n == 0:
                offset_list = list([mean_x-9.8, mean_y-0, mean_z-0])
            elif n == 1 :
                offset_list = list([mean_x+9.8, mean_y-0, mean_z-0])
            elif n == 2 :
                offset_list = list([mean_x-0, mean_y-9.8, mean_z-0])
            elif n == 3 :
                offset_list = list([mean_x-0, mean_y+9.8, mean_z-0])
            elif n == 4 :
                offset_list = list([mean_x-0, mean_y-0, mean_z-9.8])
            elif n == 5 :
                offset_list = list([mean_x-0, mean_y-0, mean_z+9.8])

            #计算灵敏度
            sens_value = 'None' #初始化灵敏度，用以給正方向灵敏度赋值
            if n == 0:
                max_x1 = max_x
            elif n == 1:
                max_x0 = min_x
                sens_value = (max_x1 - max_x0) / (2*9.807)
            elif n == 2:
                max_y1 = max_y      
            elif n == 3:
                max_y0 = min_y
                sens_value = (max_y1 - max_y0) / (2*9.807)
            elif n == 4:
                max_z1 = max_z
            elif n == 5:
                max_z0 = min_z
                sens_value = (max_z1 - max_z0) / (2*9.807)
            
   
            #chip_data = [max_xyz_list, min_xyz_list, mean_xyz_list, std_xyz_list]
            #chip_data = [tuple(max_xyz_list), tuple(min_xyz_list), tuple(mean_xyz_list), tuple(std_xyz_list)]
            #组合统计值至list中
            chip_data = [max_x, max_y, max_z, min_x, min_y, min_z, mean_x, mean_y, mean_z,
                         std_x, std_y, std_z, offset_list[0], offset_list[1], offset_list[2],
                         sens_value]

            #将统计值添加至字典中
            final_data.update({chip_num + "_" + sheetnames[n]:chip_data})

            

            '''
            print("当前的处理的文件名：" + xlsx_name + '-' +  sheetnames[n])
            print("均值：", mean_xyz_list)
            print("最大值：", max_xyz_list)
            print("最小值：", min_xyz_list)
            print("均方差(mg)：", std_xyz_list)
            '''
        #保存文件
        wb.save(xlsx_name)
        print("QMA6100:", str(xlsx_num) ,"is OK!")
    '''
    #将final_data写入txt中
    with open(".\\final_data.txt", 'w') as f:
        for key in final_data:
            #f.write('\n')
            f.writelines('"' + str(key) + '": ' + str(final_data[key]))
            f.write('\n')
        f.write('\n')
    '''
    pf = pd.DataFrame(final_data)
    #转置单元格
    #pf = pf.transpose()
    file_path = pd.ExcelWriter('Final_data.xlsx')  # 打开excel文件
    # 替换空单元格
    pf.fillna(' ', inplace=True, axis=1)
    # 输出
    pf.to_excel(file_path, encoding='utf-8', index=False,sheet_name="data")
    file_path.save()

    end = time.time()
    print("程序已运行完毕，总共用时：",end-start,"s")
    

